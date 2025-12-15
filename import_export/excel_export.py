"""
Excel export functionality
تصدير إلى Excel
"""

import io
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export data to Excel with Arabic support"""
    
    @staticmethod
    def export_to_excel(data: List[Dict], headers: Dict[str, str],
                       title: str = '', sheet_name: str = 'Sheet1') -> bytes:
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries with data
            headers: Dict mapping field names to header labels (Arabic)
            title: Report title
            sheet_name: Name of the sheet
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Set RTL for Arabic
        ws.sheet_view.rightToLeft = True
        
        row_num = 1
        
        # Add title if provided
        if title:
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            row_num = 2
        
        # Add timestamp
        timestamp = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells(f'A{row_num}:{get_column_letter(len(headers))}{row_num}')
        timestamp_cell = ws[f'A{row_num}']
        timestamp_cell.value = timestamp
        timestamp_cell.font = Font(size=10, italic=True)
        timestamp_cell.alignment = Alignment(horizontal='center')
        row_num += 2
        
        # Add headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, (field, header_text) in enumerate(headers.items(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header_text
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        row_num += 1
        
        # Add data
        for row_data in data:
            for col_num, field in enumerate(headers.keys(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = row_data.get(field, '')
                
                # Format numbers
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='right')
            
            row_num += 1
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def export_to_csv(data: List[Dict], headers: Dict[str, str]) -> str:
        """
        Export data to CSV with UTF-8 BOM
        
        Args:
            data: List of dictionaries with data
            headers: Dict mapping field names to header labels
            
        Returns:
            CSV content as string
        """
        import csv
        import io
        
        output = io.StringIO()
        
        # Write BOM for UTF-8
        output.write('\ufeff')
        
        writer = csv.DictWriter(output, fieldnames=headers.keys(), extrasaction='ignore')
        
        # Write headers
        writer.writerow(headers)
        
        # Write data
        for row_data in data:
            writer.writerow(row_data)
        
        return output.getvalue()
    
    @staticmethod
    def save_excel_file(filepath: str, data: List[Dict],
                       headers: Dict[str, str], title: str = '',
                       sheet_name: str = 'Sheet1'):
        """
        Save data to Excel file
        
        Args:
            filepath: Path to save the file
            data: List of dictionaries with data
            headers: Dict mapping field names to header labels
            title: Report title
            sheet_name: Name of the sheet
        """
        excel_bytes = ExcelExporter.export_to_excel(
            data, headers, title, sheet_name
        )
        
        with open(filepath, 'wb') as f:
            f.write(excel_bytes)
    
    @staticmethod
    def save_csv_file(filepath: str, data: List[Dict], headers: Dict[str, str]):
        """
        Save data to CSV file
        
        Args:
            filepath: Path to save the file
            data: List of dictionaries with data
            headers: Dict mapping field names to header labels
        """
        csv_content = ExcelExporter.export_to_csv(data, headers)
        
        with open(filepath, 'w', encoding='utf-8-sig') as f:
            f.write(csv_content)
